
from functools import wraps
import pymongo

from flask import jsonify

from itsdangerous import (TimedJSONWebSignatureSerializer as Serializer)
from flask_restful import reqparse
from datetime import datetime
import requests
from flask import request, url_for, Flask
from openpyxl import load_workbook
from werkzeug.contrib.atom import AtomFeed
from dicttoxml import dicttoxml
from flask_cors import *






url1 = "https://docs.google.com/spreadsheets/d/1tHCxouhyM4edDvF60VG7nzs5QxID3ADwr3DGJh71qFg/export?format=xlsx&id=1tHCxouhyM4edDvF60VG7nzs5QxID3ADwr3DGJh71qFg"
r = requests.get(url1)
postcode_dic = []

with open("temp_postcode.xlsx", "wb") as code:
    code.write(r.content)
wb = load_workbook(filename=r'temp_postcode.xlsx')
sheet = wb.get_sheet_by_name('lga_postcode_mappings')
for i in range(2,1783):
    postcode_dic.append((int(sheet.cell(row=i, column=3).value) , sheet.cell(row=i, column=2).value))



app = Flask(__name__)
CORS(app, supports_credentials=True)

SECRET_KEY = "A RANDOM KEY"


def authenticate_by_token(token):
    if token is None:
        return False
    s = Serializer(SECRET_KEY)
    try:
        username = s.loads(token.encode())
        if username == 'admin':
            return True
    except:
        return False

    return False
def authenticate_by_token_guestoradmin(token):
    if token is None:
        return False
    s = Serializer(SECRET_KEY)
    try:
        username = s.loads(token.encode())
        if username == 'guest' or username == 'admin':
            return True
    except:
        return False

    return False
def login_required_guestoradmin(f, message="You are not authorized"):
    @wraps(f)
    def decorated_function_a(*args, **kwargs):
        token = request.headers.get("AUTH_TOKEN")
        if authenticate_by_token_guestoradmin(token):
            return f(*args, **kwargs)

        return jsonify(message=message), 401
      # abort(401, message=message)

    return decorated_function_a

def login_required_admin(f, message="You are not authorized"):
    @wraps(f)
    def decorated_function(*args, **kwargs):

        token = request.headers.get("AUTH_TOKEN")
        if authenticate_by_token(token):
            return f(*args, **kwargs)

        return jsonify(message=message), 401
        # abort(401, message=message)

    return decorated_function



@app.route("/import", methods=['POST'])
@login_required_admin
def import_data():

    parser = reqparse.RequestParser()
    parser.add_argument('postcode', type=int)
    parser.add_argument('lga', type=str)
    args = parser.parse_args()

    postcode = args.get("postcode")
    lga = args.get("lga")
    feed = AtomFeed('Create', id=url_for('import_data', _external=True))
    lganame_list = []
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
      lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
      collections)
    collections = list(collections)
    if postcode:
        tt = filter(lambda t: t[0] == int(postcode), postcode_dic)
        tt = list(tt)
        for x in range(len(tt)):
            lganame_list.append(tt[x][1])
    else:
        lganame_list.append(lga)
    status = False
    if lganame_list:
        for lganame in lganame_list:
            if lganame == None:
                return jsonify('wrong lganame'), 404

            if ' ' in lganame:
                x_list = lganame.split(' ')
                lganame = ''.join(x_list)

            if lganame in collections:
                continue
            status = True
            posts = db[lganame]
            url = 'http://www.bocsar.nsw.gov.au/Documents/RCS-Annual/' + lganame + 'LGA.xlsx'
            temp_file = requests.get(url)
            if temp_file.status_code != 200:
                return jsonify('wrong lganame'), 404
            with open("temp_LGA.xlsx", "wb") as code:
                code.write(temp_file.content)
            wb = load_workbook(filename=r'temp_LGA.xlsx')
            sheet = wb.get_sheet_by_name('Summary of offences')
            result = {}
            result[lganame] = {}
            result[lganame]['2012'] = {}
            result[lganame]['2013'] = {}
            result[lganame]['2014'] = {}
            result[lganame]['2015'] = {}
            result[lganame]['2016'] = {}
            result[lganame]['trend'] = {}
            off_type_2012 = {}
            off_type_2013 = {}
            off_type_2014 = {}
            off_type_2015 = {}
            off_type_2016 = {}
            tite = sheet.cell(row=8, column=1).value
            for i in range(8, 70):
                first_trend = {}
                second_trend = {}
                rank = {}
                if sheet.cell(row=i, column=1).value != None and i != 8:
                    result[lganame]['2012'][tite] = off_type_2012
                    result[lganame]['2013'][tite] = off_type_2013
                    result[lganame]['2014'][tite] = off_type_2014
                    result[lganame]['2015'][tite] = off_type_2015
                    result[lganame]['2016'][tite] = off_type_2016
                    tite = sheet.cell(row=i, column=1).value
                    off_type_2012 = {}
                    off_type_2013 = {}
                    off_type_2014 = {}
                    off_type_2015 = {}
                    off_type_2016 = {}
                temp_2012 = {}
                temp_2013 = {}
                temp_2014 = {}
                temp_2015 = {}
                temp_2016 = {}
                temp_2012["number of incidients"] = sheet.cell(row=i, column=3).value
                temp_2012["rate per 100,000 population"] = sheet.cell(row=i, column=4).value
                temp_2013["number of incidients"] = sheet.cell(row=i, column=5).value
                temp_2013["rate per 100,000 population"] = sheet.cell(row=i, column=6).value
                temp_2014["number of incidients"] = sheet.cell(row=i, column=7).value
                temp_2014["rate per 100,000 population"] = sheet.cell(row=i, column=8).value
                temp_2015["number of incidients"] = sheet.cell(row=i, column=9).value
                temp_2015["rate per 100,000 population"] = sheet.cell(row=i, column=10).value
                temp_2016["number of incidients"] = sheet.cell(row=i, column=11).value
                temp_2016["rate per 100,000 population"] = sheet.cell(row=i, column=12).value
                first_trend["24 month trend"] = sheet.cell(row=i, column=13).value
                second_trend["60 month trend"] = sheet.cell(row=i, column=14).value
                rank["rank"] = sheet.cell(row=i, column=15).value
                if sheet.cell(row=i, column=2).value != None:
                    off_type_2012[sheet.cell(row=i, column=2).value] = temp_2012
                    off_type_2013[sheet.cell(row=i, column=2).value] = temp_2013
                    off_type_2014[sheet.cell(row=i, column=2).value] = temp_2014
                    off_type_2015[sheet.cell(row=i, column=2).value] = temp_2015
                    off_type_2016[sheet.cell(row=i, column=2).value] = temp_2016
                    result[lganame]['trend'][sheet.cell(row=i, column=2).value] = [first_trend, second_trend, rank]
                else:
                    off_type_2012 = temp_2012
                    off_type_2013 = temp_2013
                    off_type_2014 = temp_2014
                    off_type_2015 = temp_2015
                    off_type_2016 = temp_2016
                    result[lganame]['trend'][sheet.cell(row=i, column=1).value] = [first_trend, second_trend, rank]
                if i == 69:
                    result[lganame]['2012'][tite] = off_type_2012
                    result[lganame]['2013'][tite] = off_type_2013
                    result[lganame]['2014'][tite] = off_type_2014
                    result[lganame]['2015'][tite] = off_type_2015
                    result[lganame]['2016'][tite] = off_type_2016
            result = posts.insert_one(result)



            feed.add('import',content='==PLACE_HOLDER==', content_type='application/xml',
                     id=url_for('collect_information', lga_id=lganame, _external=True),
                     updated=datetime.now(),
                     author='Yijiang')

        feedstring = feed.to_string();
        if status == False:
            return jsonify('all entry already in database'), 200

        return jsonify(feedstring), 201

@app.route("/collectjson/<string:lga_id>", methods=["GET"])
@login_required_guestoradmin
def collect_information_json(lga_id):

    if lga_id.isdigit():
        lganame = postcode_dic[int(lga_id)]
    else:
        lganame = lga_id
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
      lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
      collections)
    collections = list(collections)
    if lganame not in collections:
        return jsonify('wrong lga_id'), 404
    temp = db[lganame].find()[0][lganame]

    return jsonify(temp), 200

@app.route("/collectxml/<string:lga_id>", methods=["GET"])
@login_required_guestoradmin
def collect_information(lga_id):

    if lga_id.isdigit():
        lganame = postcode_dic[int(lga_id)]
    else:
        lganame = lga_id
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
      lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
      collections)
    collections = list(collections)
    if lganame not in collections:
        return jsonify('wrong lga_id'), 404
    temp = db[lganame].find()[0][lganame]
    feed = AtomFeed('Retreival result', id=url_for('collect_information', lga_id=lganame, _external=True))

    feed.add('get one entry',content='==PLACE_HOLDER==', content_type='application/xml',
             id=url_for('collect_information', lga_id=lganame, _external=True),
             updated=datetime.now(),
             author='Yijiang')

    feedstring = feed.to_string().replace('==PLACE_HOLDER==',
                                      dicttoxml(temp, root=False).decode())

    return jsonify(feedstring), 200

@app.route("/collects", methods=['GET'])
@login_required_guestoradmin
def collect_allinformation():
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
        lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
        collections)
    collections = list(collections)
    feed = AtomFeed('all collections result', id=url_for('collect_allinformation', _external=True))
    for collection in collections:

        # temp = db[collection].find()[0][collection]
        feed.add('entry',content = '==PLACE_HOLDER==', content_type='application/xml',
                 id=url_for('collect_information', lga_id=collection, _external=True),
                 updated=datetime.now(),
                 author='Yijiang')
    feedstring = feed.to_string()
    return jsonify(feedstring), 200

@app.route("/collectsjson", methods=['GET'])
@login_required_guestoradmin
def collect_allinformation_json():
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
        lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
        collections)
    collections = list(collections)
    result = {}
    for collection in collections:

        temp = db[collection].find()[0][collection]
        result[collection] = temp

    return jsonify(result), 200
@app.route("/filter", methods=['POST'])
@login_required_guestoradmin
def collection_filter():
    parser = reqparse.RequestParser()
    parser.add_argument('year', type=str)
    parser.add_argument('lga1', type=str)
    parser.add_argument('lga2', type=str)
    args = parser.parse_args()

    year = args.get("year")
    lga1 = args.get("lga1")
    lga2 = args.get("lga2")
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
      lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
      collections)
    collections = list(collections)
    if year:
        if lga1 in collections:
            temp = db[lga1].find()[0][lga1][year]
            return jsonify(temp), 200
        else:
            return jsonify(lga1=False), 404
    else:
        if lga2 in collections and lga1 in collections:
            temp1 = db[lga1].find()[0][lga1]
            temp2 = db[lga2].find()[0][lga2]
            temp = {}
            temp[lga1] = temp1
            temp[lga2] = temp2
            return jsonify(temp), 200
        else:
            return jsonify('wrong lga1 or lga2'), 404

@app.route("/delete", methods=['DELETE'])
@login_required_admin
def delete_entry():
    parser = reqparse.RequestParser()
    parser.add_argument('lga', type=str)
    args = parser.parse_args()
    lga = args.get("lga")
    if lga == None:
        return jsonify('wrong lganame'), 404
    client = pymongo.MongoClient('mongodb://wyj199481:wyj112781@ds255889.mlab.com:55889/assignment2')
    db = client.assignment2
    db.authenticate("wyj199481", "wyj112781")
    collections = db.collection_names()
    collections = filter(
      lambda t: t != 'objectlabs-system.admin.collections' and t != 'system.indexes' and t != 'objectlabs-system',
      collections)
    collections = list(collections)
    if lga not in collections:
        return jsonify('wrong lga_id'), 404
    db[lga].drop()
    return jsonify('delete successful'), 200


@app.route("/auth", methods=['POST'])
def generate_token():
        parser = reqparse.RequestParser()
        parser.add_argument('username', type=str)
        parser.add_argument('password', type=str)
        args = parser.parse_args()
        username = args.get('username')
        password = args.get('password')
        print(username)
        s = Serializer(SECRET_KEY, expires_in=600)
        token = s.dumps(username)

        if (username == 'admin' and password == 'admin') or(username == 'guest' and password == 'guest'):
            # return token.decode()
            response = jsonify(token.decode())
            # response.headers._list.append(('Access-Control-Allow-Origin', '*'))
            return response, 200

        return 404


if __name__ == "__main__":
    app.run()
